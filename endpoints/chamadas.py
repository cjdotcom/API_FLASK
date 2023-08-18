from flask_restx import Resource, Api
from flask import request
from server.instance import server
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
import json

app, api = server.app, server.api

@api.route('/Status')
class ApiStatus(Resource):
    def get(self,):
        dia = datetime.now() #data atual
        lista = """{
            "Status": "Okay",
            "Date": ""
            }"""
        db = json.loads(lista)
        db['Date'] = str(dia)
        return db

@api.route('/Products')
class Produtos(Resource):
    def get(self,):
        caminhoExcel = "db_excel.xlsx"
        wb = load_workbook(caminhoExcel)
        sheet = wb['DADOS']
        lista = []
        for i in range(2, sheet.max_row+1):
            linha = [x.value for x in sheet[i]]
            if linha[0] != None:
                lista.append(linha)
            else:
                break

        t = {}
        for i in lista:
            t[i[0]] = [i[1], i[2]]

        form = """
            {
            "codigo":"",
            "name":"",
            "dtRegistro":""
            }
        """
        tdb = json.loads(form)

        retorno = []
        for k, v in t.items():
            info = tdb.copy()

            info['codigo'] = k
            info['name'] = v[0]
            info['dtRegistro'] = v[1]

            retorno.append(info)

        products ={
            "Products":""
        }
        products['Products'] = retorno

        result = json.dumps(products, indent=2)
        return json.loads(result)
    
@api.route('/Product')
class ProdutoByCodigo(Resource):
    def get(self,):
        codigo = request.args.get('codigo')
        caminhoExcel = "db_excel.xlsx"
        wb = load_workbook(caminhoExcel)
        sheet = wb['DADOS']
        lista = []
        for i in range(2, sheet.max_row+1):
            linha = [x.value for x in sheet[i]]
            if linha[0] != None:
                lista.append(linha)
            else:
                break

        t = {}
        for i in lista:
            t[i[0]] = [i[1], i[2]]

        form = """
            {
            "codigo":"",
            "name":"",
            "dtRegistro":""
            }
        """
        tdb = json.loads(form)

        retorno = []
        for k, v in t.items():
            info = tdb.copy()

            info['codigo'] = k
            info['name'] = v[0]
            info['dtRegistro'] = v[1]

            retorno.append({"Products":info})

        aviso = {
            "erro": "",
            "param": ""
            }
            
        aviso['erro'] = f"Essa chamada requer o parametro <codigo>"
        aviso['param'] = "missing"
        jsonAviso = json.dumps(aviso, indent=2)

        if codigo != None:
            df = pd.DataFrame(retorno)
            df_real = None
            for db_cod in df['Products'].values:
                if db_cod['codigo'] == int(codigo):
                    df_real = json.dumps({"Product":db_cod})
            
            if df_real != None:
                return json.loads(df_real)
            else:
                aviso = {
                        "erro": ""
                    }
                
                aviso['erro'] = f"Código '{codigo}' não cadastrado!"
                jsonAviso = json.dumps(aviso, indent=2)
                return json.loads(jsonAviso)
        else:
            return json.loads(jsonAviso)
        
    def post(self,):
        codigo = request.args['codigo']
        nomeProduto = request.args['nomeProduto']

        caminhoExcel = "db_excel.xlsx"
        wb = load_workbook(caminhoExcel)
        sheet = wb['DADOS']
        lista = []

        c = 1
        for i in range(2, sheet.max_row+2):
            linha = [x.value for x in sheet[i]]
            c += 1
            if linha[0] != None:
                lista.append(linha)
            else:
                break

        t = {}
        for i in lista:
            t[i[0]] = [i[1], i[2]]

        gatilho = None
        for k, v in t.items():
            if int(codigo) == k:
                gatilho = True
                break
            else:
                gatilho = False

        if gatilho == False:
            linha = c
            sheet.cell(row=linha, column=1, value=int(codigo))
            sheet.cell(row=linha, column=2, value=str(nomeProduto))
            sheet.cell(row=linha, column=3, value=datetime.today().strftime("%Y-%m-%d %H:%M:%S"))

            wb.save(caminhoExcel)

            aviso = {
                    "info":""
                }
            aviso['info'] = f"Produto ({codigo}, {nomeProduto}) cadastrado!"
        else:
            aviso = {
                    "error":"Produto "
                }
            aviso['error'] = f"Produto {codigo} já existe."

        jsonAviso = json.dumps(aviso, indent=2)

        return json.loads(jsonAviso)
        
    def delete(self,):
        codigo = request.args['codigo']
        caminhoExcel = "db_excel.xlsx"
        wb = load_workbook(caminhoExcel)
        sheet = wb['DADOS']

        index = None
        for i, linha in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True), start=1):
            if linha[0] == codigo:
                index = i
                break
        
        if index is not None:
            sheet.delete_rows(index)
            aviso = {
                    "Succes":""
                }
            aviso['Succes'] = f"Produto {codigo} excluído!"
        else:
            aviso = {
                "Error":""
            }
            aviso['Error'] = f"Produto {codigo} não existe."
        jsonAviso = json.dumps(aviso, indent=2)

        return json.loads(jsonAviso)