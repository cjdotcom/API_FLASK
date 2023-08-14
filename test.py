import pandas as pd
import json
from openpyxl import load_workbook
from datetime import datetime
from pprint import pprint


# Pega dados da planilha
def getProdutos(codigo=0):
    caminhoExcel = "db_excel.xlsx"
    wb = load_workbook(caminhoExcel)
    sheet = wb['DADOS']
    lista = []
    for i in range(2, sheet.max_row):
        linha = [x.value for x in sheet[i]]
        if linha[0] != None:
            lista.append(linha)
        else:
            break

    t = {}
    for i in lista:
        t[i[0]] = [i[1], i[2]]

    form = """
        {
        "codigo":"",
        "name":"",
        "dtRegistro":""
        }
    """
    tdb = json.loads(form)

    retorno = []
    for k, v in t.items():
        info = tdb.copy()

        info['codigo'] = k
        info['name'] = v[0]
        info['dtRegistro'] = v[1]

        retorno.append({"Products":info})

    df = pd.DataFrame(retorno)
    df_real = None
    for db_cod in df['Products'].values:
        if db_cod['codigo'] == codigo:
            df_real = json.dumps({"Product":db_cod})
    
    if df_real != None:
        return json.loads(df_real)
    else:
        erro = """
        {
            "erro": "Código não cadastrado!"
        }
        """
        return json.loads(erro)

    # return json.loads(df.to_json())

# pprint(getProdutos(10))

#Escreve informações na planilha
def postProduto(codigo, nomeproduto):
    caminhoExcel = 'db_excel.xlsx'
    workbook = load_workbook(caminhoExcel)
    sheet = workbook['DADOS']
    c = 0
    for i in range(2, sheet.max_row):
        linha = [x.value for x in sheet[i]]
        if linha[0] != None:
            c += 1
        else:
            break

    
    dados = getProdutos()
    gatilho = None
    for k, v in dados.items():
        if codigo == k:
            print("Código já existe")
            gatilho = True
            break
        else:
            gatilho = False

    if gatilho == False:
        linha = c+1
        print(f"Linha de inserção: {linha}")
        sheet.cell(row=linha, column=1, value=int(codigo))
        sheet.cell(row=linha, column=2, value=nomeproduto)
        sheet.cell(row=linha, column=3, value=datetime.today().strftime("%d/%m/%Y %H:%M:%S"))

        workbook.save(caminhoExcel)
            #Roda escrita do código na planilha

# postProduto(1005, 'produto10')

# form = """{
#   "Products": [
#     {
#       "codigo": 1000,
#       "nome": "produto 1",
#       "dataCriacao": "00/00/0000"
#     }
#   ]
# }
# """

# # with open("dados.json", "w") as criar:
# #     criar.write(form)

# with open("dados.json", "a") as adicionar:
#     adicionar.writelines()